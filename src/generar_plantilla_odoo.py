#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Genera una plantilla robusta para importación manual de Traslados Internos en Odoo.

Uso en GitHub Actions:
    python src/generar_plantilla_odoo.py generar --origen "CUATI/Existencias" --salida "salidas"

Secrets requeridos en GitHub:
    ODOO_URL
    ODOO_DB
    ODOO_USER
    ODOO_API_KEY

Qué genera:
    - LEEME
    - CAPTURA
    - IMPORT_ODOO_IDS
    - CAT_STOCK_CUATI
    - CAT_UBICACIONES
    - CAT_CONTACTOS
    - CONFIG
"""

from __future__ import annotations

import argparse
import os
import sys
import xmlrpc.client
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


# =============================================================================
# CONFIGURACIÓN
# =============================================================================

DEFAULT_SOURCE_LOCATION_NAME = "CUATI/Existencias"
DEFAULT_OUTPUT_DIR = Path("salidas")
MAX_CAPTURE_ROWS = 1000

SH_README = "LEEME"
SH_CAPTURE = "CAPTURA"
SH_IMPORT = "IMPORT_ODOO_IDS"
SH_STOCK = "CAT_STOCK_CUATI"
SH_LOC = "CAT_UBICACIONES"
SH_PARTNERS = "CAT_CONTACTOS"
SH_CONFIG = "CONFIG"

CAPTURE_HEADERS = [
    "Cargar",
    "Documento origen",
    "Contacto",
    "Ubicación de origen",
    "Ubicación de destino",
    "Producto + lote disponible",
    "Cantidad a mover",
    "Comentarios",
    "Stock disponible",
    "Producto",
    "Lote",
    "product_id",
    "lot_id",
    "source_location_id",
    "dest_location_id",
    "partner_id",
]

ODOO_IMPORT_HEADERS = [
    "ID",
    "Documento origen",
    "Contacto/ID de la base de datos",
    "Ubicación de origen/ID de la base de datos",
    "Ubicación de destino/ID de la base de datos",
    "Operaciones/Producto/ID de la base de datos",
    "Operaciones/Número de serie/lote/ID de la base de datos",
    "Operaciones/Cantidad",
    "Operaciones/Desde/ID de la base de datos",
    "Operaciones/A/ID de la base de datos",
    "Nota",
]

STOCK_HEADERS = [
    "item_key",
    "product_id",
    "product_display_name",
    "default_code",
    "lot_id",
    "lot_name",
    "location_id",
    "location_complete_name",
    "quantity",
    "reserved_quantity",
    "available_quantity",
    "tracking",
]

LOC_HEADERS = ["location_id", "complete_name", "name", "usage", "active"]
PARTNER_HEADERS = ["partner_id", "display_name", "name", "email", "active"]


# =============================================================================
# ESTILOS
# =============================================================================

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LOCKED_FILL = PatternFill("solid", fgColor="E7E6E6")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9E1F2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# =============================================================================
# CLIENTE ODOO
# =============================================================================

class OdooClient:
    def __init__(self) -> None:
        self.url = os.getenv("ODOO_URL", "").strip().rstrip("/")
        self.db = os.getenv("ODOO_DB", "").strip()
        self.user = os.getenv("ODOO_USER", "").strip()
        self.api_key = os.getenv("ODOO_API_KEY", "").strip()

        if not all([self.url, self.db, self.user, self.api_key]):
            raise RuntimeError(
                "Faltan secrets de Odoo. Configura ODOO_URL, ODOO_DB, "
                "ODOO_USER y ODOO_API_KEY en GitHub Secrets."
            )

        self.common = xmlrpc.client.ServerProxy(
            f"{self.url}/xmlrpc/2/common",
            allow_none=True,
        )
        self.models = xmlrpc.client.ServerProxy(
            f"{self.url}/xmlrpc/2/object",
            allow_none=True,
        )
        self.uid: Optional[int] = None

    def connect(self) -> int:
        uid = self.common.authenticate(self.db, self.user, self.api_key, {})
        if not uid:
            raise RuntimeError(
                "No se pudo autenticar en Odoo. Revisa URL, DB, usuario y API key."
            )
        self.uid = int(uid)
        print(f"Conectado a Odoo. UID: {self.uid}")
        return self.uid

    def execute_kw(
        self,
        model: str,
        method: str,
        args: list,
        kwargs: Optional[dict] = None,
    ) -> Any:
        if self.uid is None:
            raise RuntimeError("Cliente no conectado. Ejecuta connect() primero.")

        return self.models.execute_kw(
            self.db,
            self.uid,
            self.api_key,
            model,
            method,
            args,
            kwargs or {},
        )

    def fields_get(self, model: str) -> Dict[str, Any]:
        return self.execute_kw(
            model,
            "fields_get",
            [],
            {"attributes": ["string", "type", "store"]},
        )

    def search_read(
        self,
        model: str,
        domain: list,
        fields: list,
        limit: Optional[int] = None,
        offset: int = 0,
        order: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        kwargs: Dict[str, Any] = {
            "fields": fields,
            "offset": offset,
        }

        if limit is not None:
            kwargs["limit"] = limit

        if order:
            kwargs["order"] = order

        return self.execute_kw(model, "search_read", [domain], kwargs)

    def search_read_all(
        self,
        model: str,
        domain: list,
        fields: list,
        batch_size: int = 1000,
        order: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        offset = 0

        while True:
            chunk = self.search_read(
                model=model,
                domain=domain,
                fields=fields,
                limit=batch_size,
                offset=offset,
                order=order,
            )

            if not chunk:
                break

            rows.extend(chunk)

            if len(chunk) < batch_size:
                break

            offset += batch_size

        return rows


# =============================================================================
# UTILIDADES
# =============================================================================

def m2o_id(value: Any) -> Optional[int]:
    if isinstance(value, (list, tuple)) and value:
        return int(value[0]) if value[0] else None
    if isinstance(value, int):
        return value
    return None


def m2o_name(value: Any) -> str:
    if isinstance(value, (list, tuple)) and len(value) >= 2:
        return str(value[1] or "")
    return ""


def safe_float(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def make_odoo_external_id(*parts: Any) -> str:
    raw = "_".join(str(p or "") for p in parts)
    raw = raw.lower().strip()

    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "/": "_",
        "\\": "_",
        " ": "_",
        "-": "_",
        ".": "_",
        ",": "_",
        ":": "_",
        ";": "_",
        "(": "",
        ")": "",
        "[": "",
        "]": "",
    }

    for old, new in replacements.items():
        raw = raw.replace(old, new)

    clean = "".join(c for c in raw if c.isalnum() or c == "_")

    while "__" in clean:
        clean = clean.replace("__", "_")

    clean = clean.strip("_")

    if not clean:
        clean = "traslado_sin_referencia"

    return f"traslado_{clean}"


# =============================================================================
# EXTRACCIÓN DESDE ODOO
# =============================================================================

def get_internal_locations(client: OdooClient) -> List[Dict[str, Any]]:
    print("Descargando ubicaciones internas...")

    fields = ["id", "complete_name", "name", "usage", "active"]
    domain = [["usage", "=", "internal"], ["active", "=", True]]

    locations = client.search_read_all(
        "stock.location",
        domain,
        fields,
        order="complete_name asc",
    )

    print(f"Ubicaciones internas encontradas: {len(locations)}")
    return locations


def find_location(locations: List[Dict[str, Any]], target_name: str) -> Dict[str, Any]:
    target_norm = target_name.strip().lower()

    for loc in locations:
        if str(loc.get("complete_name", "")).strip().lower() == target_norm:
            return loc

    for loc in locations:
        complete = str(loc.get("complete_name", "")).strip().lower()
        if target_norm in complete or complete.endswith(target_norm):
            return loc

    available = "\n".join(
        f"- {l.get('complete_name')} | ID {l.get('id')}"
        for l in locations[:50]
    )

    raise RuntimeError(
        f"No encontré la ubicación origen '{target_name}'.\n"
        f"Ubicaciones disponibles:\n{available}"
    )


def get_partners(client: OdooClient) -> List[Dict[str, Any]]:
    print("Descargando contactos...")

    fields = ["id", "display_name", "name", "email", "active"]
    domain = [["active", "=", True]]

    try:
        partners = client.search_read_all(
            "res.partner",
            domain,
            fields,
            order="name asc",
        )
    except Exception as exc:
        print("Aviso: no pude ordenar contactos por name. Reintentando sin orden.")
        print(f"Detalle: {exc}")
        partners = client.search_read_all("res.partner", domain, fields)

    print(f"Contactos encontrados: {len(partners)}")
    return partners


def get_stock_quants_for_location(
    client: OdooClient,
    source_location_id: int,
) -> List[Dict[str, Any]]:
    print(f"Descargando stock.quant para ubicación ID {source_location_id}...")

    stock_fields_meta = client.fields_get("stock.quant")

    fields = [
        "id",
        "product_id",
        "location_id",
        "quantity",
        "reserved_quantity",
    ]

    if "available_quantity" in stock_fields_meta:
        fields.append("available_quantity")

    if "lot_id" in stock_fields_meta:
        fields.append("lot_id")

    domain = [
        ["location_id", "=", source_location_id],
        ["quantity", ">", 0],
    ]

    quants = client.search_read_all(
        "stock.quant",
        domain,
        fields,
        order="product_id asc",
    )

    print(f"Quants encontrados en origen: {len(quants)}")
    return quants


def get_products_by_ids(
    client: OdooClient,
    product_ids: List[int],
) -> Dict[int, Dict[str, Any]]:
    if not product_ids:
        return {}

    print(f"Descargando detalle de productos: {len(product_ids)}...")

    meta = client.fields_get("product.product")

    fields = ["id", "display_name", "name", "default_code"]

    if "tracking" in meta:
        fields.append("tracking")

    products = client.search_read_all(
        "product.product",
        [["id", "in", product_ids]],
        fields,
        order="default_code asc",
    )

    return {int(p["id"]): p for p in products}


def build_stock_rows(
    client: OdooClient,
    source_location: Dict[str, Any],
) -> List[Dict[str, Any]]:
    source_location_id = int(source_location["id"])
    source_location_name = str(
        source_location.get("complete_name")
        or source_location.get("name")
        or ""
    )

    quants = get_stock_quants_for_location(client, source_location_id)

    product_ids = sorted({
        m2o_id(q.get("product_id"))
        for q in quants
        if m2o_id(q.get("product_id"))
    })

    products_by_id = get_products_by_ids(client, product_ids)  # type: ignore[arg-type]

    grouped: Dict[Tuple[int, int, int], Dict[str, Any]] = {}

    for q in quants:
        product_id = m2o_id(q.get("product_id"))

        if not product_id:
            continue

        lot_id = m2o_id(q.get("lot_id")) or 0
        location_id = m2o_id(q.get("location_id")) or source_location_id

        key = (product_id, lot_id, location_id)

        quantity = safe_float(q.get("quantity"))
        reserved = safe_float(q.get("reserved_quantity"))

        if "available_quantity" in q:
            available = safe_float(q.get("available_quantity"))
        else:
            available = quantity - reserved

        if key not in grouped:
            product = products_by_id.get(product_id, {})
            product_display = str(
                product.get("display_name")
                or m2o_name(q.get("product_id"))
            )
            default_code = str(product.get("default_code") or "")
            tracking = str(product.get("tracking") or "")
            lot_name = m2o_name(q.get("lot_id")) if lot_id else "SIN LOTE"

            grouped[key] = {
                "item_key": "",
                "product_id": product_id,
                "product_display_name": product_display,
                "default_code": default_code,
                "lot_id": lot_id,
                "lot_name": lot_name,
                "location_id": location_id,
                "location_complete_name": source_location_name,
                "quantity": 0.0,
                "reserved_quantity": 0.0,
                "available_quantity": 0.0,
                "tracking": tracking,
            }

        grouped[key]["quantity"] += quantity
        grouped[key]["reserved_quantity"] += reserved
        grouped[key]["available_quantity"] += available

    rows = [
        r for r in grouped.values()
        if safe_float(r["available_quantity"]) > 0
    ]

    rows.sort(
        key=lambda r: (
            str(r["default_code"]),
            str(r["product_display_name"]),
            str(r["lot_name"]),
        )
    )

    for r in rows:
        default_code = str(r.get("default_code") or "").strip()
        product_display = str(r["product_display_name"])
        lot_name = str(r["lot_name"])
        available = safe_float(r["available_quantity"])

        if default_code:
            r["item_key"] = (
                f"[{default_code}] {product_display} | "
                f"Lote: {lot_name} | Disp: {available:g}"
            )
        else:
            r["item_key"] = (
                f"{product_display} | "
                f"Lote: {lot_name} | Disp: {available:g}"
            )

    print(f"Renglones producto+lote disponibles: {len(rows)}")
    return rows


# =============================================================================
# EXCEL
# =============================================================================

def apply_basic_style(
    ws,
    max_row: int,
    max_col: int,
    header_row: int = 1,
) -> None:
    for cell in ws[header_row]:
        if cell.column <= max_col:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = BORDER

    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def set_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_row(ws, row_idx: int, values: List[Any]) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


def create_readme(wb: Workbook) -> None:
    ws = wb.create_sheet(SH_README)

    ws["A1"] = "Plantilla de traslados internos Odoo - importación por IDs"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")

    ws["A3"] = "Objetivo"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = (
        "Generar una plantilla para mover stock desde CUATI/Existencias "
        "a cualquier ubicación interna disponible en Odoo."
    )

    instructions = [
        "1. Llena únicamente la hoja CAPTURA.",
        "2. En CAPTURA, selecciona Producto + lote disponible desde la lista desplegable.",
        "3. Selecciona la ubicación de destino.",
        "4. Captura la cantidad a mover.",
        "5. No modifiques las columnas grises; contienen IDs internos para Odoo.",
        "6. Para subir a Odoo, usa la hoja IMPORT_ODOO_IDS.",
        "7. En Odoo, mapea los campos con '/ID de la base de datos'.",
        "8. En 'Cuando un valor no puede coincidir', usa 'Prevenir importación'.",
        "9. Odoo generará automáticamente la referencia oficial del traslado.",
    ]

    for i, text in enumerate(instructions, start=5):
        ws.cell(i, 1, text)

    ws["A16"] = "Campos clave para importar"
    ws["A16"].font = Font(bold=True)

    for i, h in enumerate(ODOO_IMPORT_HEADERS, start=17):
        ws.cell(i, 1, h)

    set_widths(ws, {"A": 55, "B": 100})


def create_catalog_sheets(
    wb: Workbook,
    locations: List[Dict[str, Any]],
    partners: List[Dict[str, Any]],
    stock_rows: List[Dict[str, Any]],
) -> None:
    # Ubicaciones
    ws = wb.create_sheet(SH_LOC)
    write_row(ws, 1, LOC_HEADERS)

    for r, loc in enumerate(locations, start=2):
        write_row(ws, r, [
            loc.get("id"),
            loc.get("complete_name"),
            loc.get("name"),
            loc.get("usage"),
            loc.get("active"),
        ])

    apply_basic_style(ws, max_row=max(2, len(locations) + 1), max_col=len(LOC_HEADERS))
    set_widths(ws, {"A": 12, "B": 38, "C": 28, "D": 14, "E": 12})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(2, len(locations) + 1)}"

    # Contactos
    ws = wb.create_sheet(SH_PARTNERS)
    write_row(ws, 1, PARTNER_HEADERS)

    for r, p in enumerate(partners, start=2):
        write_row(ws, r, [
            p.get("id"),
            p.get("display_name"),
            p.get("name"),
            p.get("email"),
            p.get("active"),
        ])

    apply_basic_style(ws, max_row=max(2, len(partners) + 1), max_col=len(PARTNER_HEADERS))
    set_widths(ws, {"A": 12, "B": 44, "C": 38, "D": 34, "E": 12})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(2, len(partners) + 1)}"

    # Stock
    ws = wb.create_sheet(SH_STOCK)
    write_row(ws, 1, STOCK_HEADERS)

    for r, row in enumerate(stock_rows, start=2):
        write_row(ws, r, [row.get(h) for h in STOCK_HEADERS])

    apply_basic_style(ws, max_row=max(2, len(stock_rows) + 1), max_col=len(STOCK_HEADERS))
    set_widths(ws, {
        "A": 95,
        "B": 12,
        "C": 48,
        "D": 18,
        "E": 12,
        "F": 24,
        "G": 12,
        "H": 34,
        "I": 12,
        "J": 16,
        "K": 18,
        "L": 14,
    })
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(2, len(stock_rows) + 1)}"


def create_config_sheet(
    wb: Workbook,
    source_location: Dict[str, Any],
) -> None:
    ws = wb.create_sheet(SH_CONFIG)

    rows = [
        ["Parámetro", "Valor"],
        ["source_location_name", source_location.get("complete_name")],
        ["source_location_id", source_location.get("id")],
        ["fecha_generacion", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["nota", "Plantilla general con todas las ubicaciones internas disponibles."],
    ]

    for i, row in enumerate(rows, start=1):
        write_row(ws, i, row)

    apply_basic_style(ws, max_row=len(rows), max_col=2)
    set_widths(ws, {"A": 30, "B": 100})


def create_capture_sheet(
    wb: Workbook,
    source_location: Dict[str, Any],
    locations_count: int,
    partners_count: int,
    stock_count: int,
) -> None:
    ws = wb.create_sheet(SH_CAPTURE)

    write_row(ws, 1, CAPTURE_HEADERS)
    apply_basic_style(
        ws,
        max_row=MAX_CAPTURE_ROWS + 1,
        max_col=len(CAPTURE_HEADERS),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{MAX_CAPTURE_ROWS + 1}"

    editable_cols = [1, 2, 3, 5, 6, 7, 8]
    locked_cols = [4, 9, 10, 11, 12, 13, 14, 15, 16]

    for row in range(2, MAX_CAPTURE_ROWS + 2):
        for col in editable_cols:
            ws.cell(row, col).fill = EDIT_FILL

        for col in locked_cols:
            ws.cell(row, col).fill = LOCKED_FILL

    source_name = source_location.get("complete_name") or source_location.get("name")
    source_id = source_location.get("id")

    for r in range(2, MAX_CAPTURE_ROWS + 2):
        ws.cell(r, 4, source_name)
        ws.cell(r, 14, source_id)

        ws.cell(r, 9, f'=IFERROR(VLOOKUP(F{r},{SH_STOCK}!$A:$L,11,FALSE),"")')
        ws.cell(r, 10, f'=IFERROR(VLOOKUP(F{r},{SH_STOCK}!$A:$L,3,FALSE),"")')
        ws.cell(r, 11, f'=IFERROR(VLOOKUP(F{r},{SH_STOCK}!$A:$L,6,FALSE),"")')
        ws.cell(r, 12, f'=IFERROR(VLOOKUP(F{r},{SH_STOCK}!$A:$L,2,FALSE),"")')
        ws.cell(r, 13, f'=IFERROR(VLOOKUP(F{r},{SH_STOCK}!$A:$L,5,FALSE),"")')
        ws.cell(r, 15, f'=IFERROR(INDEX({SH_LOC}!$A:$A,MATCH(E{r},{SH_LOC}!$B:$B,0)),"")')
        ws.cell(r, 16, f'=IFERROR(INDEX({SH_PARTNERS}!$A:$A,MATCH(C{r},{SH_PARTNERS}!$B:$B,0)),"")')

    dv_yes = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    ws.add_data_validation(dv_yes)
    dv_yes.add(f"A2:A{MAX_CAPTURE_ROWS + 1}")

    if partners_count > 0:
        dv_partner = DataValidation(
            type="list",
            formula1=f"={SH_PARTNERS}!$B$2:$B${partners_count + 1}",
            allow_blank=True,
        )
        ws.add_data_validation(dv_partner)
        dv_partner.add(f"C2:C{MAX_CAPTURE_ROWS + 1}")

    if locations_count > 0:
        dv_loc = DataValidation(
            type="list",
            formula1=f"={SH_LOC}!$B$2:$B${locations_count + 1}",
            allow_blank=False,
        )
        ws.add_data_validation(dv_loc)
        dv_loc.add(f"E2:E{MAX_CAPTURE_ROWS + 1}")

    if stock_count > 0:
        dv_stock = DataValidation(
            type="list",
            formula1=f"={SH_STOCK}!$A$2:$A${stock_count + 1}",
            allow_blank=False,
        )
        ws.add_data_validation(dv_stock)
        dv_stock.add(f"F2:F{MAX_CAPTURE_ROWS + 1}")

    set_widths(ws, {
        "A": 10,
        "B": 24,
        "C": 30,
        "D": 28,
        "E": 32,
        "F": 95,
        "G": 16,
        "H": 32,
        "I": 16,
        "J": 48,
        "K": 24,
        "L": 12,
        "M": 12,
        "N": 18,
        "O": 18,
        "P": 12,
    })

    for col in ["L", "M", "N", "O", "P"]:
        ws.column_dimensions[col].hidden = True


def create_import_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SH_IMPORT)

    write_row(ws, 1, ODOO_IMPORT_HEADERS)
    apply_basic_style(
        ws,
        max_row=MAX_CAPTURE_ROWS + 1,
        max_col=len(ODOO_IMPORT_HEADERS),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{MAX_CAPTURE_ROWS + 1}"

    for r in range(2, MAX_CAPTURE_ROWS + 2):
        cap_row = r

        ws.cell(
            r,
            1,
            f'=IF({SH_CAPTURE}!A{cap_row}="Sí",'
            f'"traslado_"&LOWER(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
            f'{SH_CAPTURE}!B{cap_row}&"_"&{SH_CAPTURE}!D{cap_row}&"_"&{SH_CAPTURE}!E{cap_row},'
            f'"/","_")," ","_"),"-","_")),"")',
        )
        ws.cell(r, 2, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!B{cap_row},"")')
        ws.cell(r, 3, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!P{cap_row},"")')
        ws.cell(r, 4, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!N{cap_row},"")')
        ws.cell(r, 5, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!O{cap_row},"")')
        ws.cell(r, 6, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!L{cap_row},"")')
        ws.cell(r, 7, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",IF({SH_CAPTURE}!M{cap_row}=0,"",{SH_CAPTURE}!M{cap_row}),"")')
        ws.cell(r, 8, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!G{cap_row},"")')
        ws.cell(r, 9, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!N{cap_row},"")')
        ws.cell(r, 10, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!O{cap_row},"")')
        ws.cell(r, 11, f'=IF({SH_CAPTURE}!A{cap_row}="Sí",{SH_CAPTURE}!H{cap_row},"")')

    set_widths(ws, {
        "A": 38,
        "B": 24,
        "C": 28,
        "D": 34,
        "E": 36,
        "F": 34,
        "G": 42,
        "H": 18,
        "I": 34,
        "J": 28,
        "K": 34,
    })


def finalize_workbook(wb: Workbook) -> None:
    if "Sheet" in wb.sheetnames:
        ws = wb["Sheet"]
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            wb.remove(ws)

    desired_order = [
        SH_README,
        SH_CAPTURE,
        SH_IMPORT,
        SH_STOCK,
        SH_LOC,
        SH_PARTNERS,
        SH_CONFIG,
    ]

    wb._sheets = [wb[s] for s in desired_order if s in wb.sheetnames]

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True


# =============================================================================
# GENERAR PLANTILLA
# =============================================================================

def generate_template(
    source_location_name: str,
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    client = OdooClient()
    client.connect()

    locations = get_internal_locations(client)
    source_location = find_location(locations, source_location_name)

    print(
        f"Ubicación origen principal: "
        f"{source_location.get('complete_name')} | ID: {source_location.get('id')}"
    )

    load_partners = os.getenv("LOAD_PARTNERS", "1").strip() != "0"

    if load_partners:
        try:
            partners = get_partners(client)
        except Exception as exc:
            print("Aviso: no pude descargar contactos. Continuaré sin contactos.")
            print(f"Detalle: {exc}")
            partners = []
    else:
        partners = []

    stock_rows = build_stock_rows(client, source_location)

    if not stock_rows:
        print("Advertencia: no se encontró stock disponible en la ubicación origen.")

    wb = Workbook()

    create_readme(wb)
    create_catalog_sheets(wb, locations, partners, stock_rows)
    create_config_sheet(wb, source_location)
    create_capture_sheet(
        wb,
        source_location=source_location,
        locations_count=len(locations),
        partners_count=len(partners),
        stock_count=len(stock_rows),
    )
    create_import_sheet(wb)
    finalize_workbook(wb)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = output_dir / f"Plantilla_Traslados_Internos_Odoo_IDs_v4_{timestamp}.xlsx"

    wb.save(output_path)

    print("\nPlantilla generada correctamente:")
    print(output_path)

    return output_path


# =============================================================================
# PREPARAR ARCHIVO FINAL PARA ODOO
# =============================================================================

def header_map(ws) -> Dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def build_lookup_from_sheet(
    ws,
    key_col_name: str,
    headers: List[str],
) -> Dict[str, Dict[str, Any]]:
    hm = header_map(ws)

    if key_col_name not in hm:
        return {}

    key_col = hm[key_col_name]
    lookup: Dict[str, Dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, key_col).value

        if key in (None, ""):
            continue

        record: Dict[str, Any] = {}

        for h in headers:
            col = hm.get(h)
            record[h] = ws.cell(row, col).value if col else None

        lookup[str(key).strip()] = record

    return lookup


def prepare_upload_file(
    template_path: Path,
    output_dir: Optional[Path] = None,
) -> Path:
    if not template_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {template_path}")

    output_dir = output_dir or template_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path, data_only=False)

    if SH_CAPTURE not in wb.sheetnames:
        raise RuntimeError(f"No existe la hoja {SH_CAPTURE} en {template_path.name}")

    if SH_STOCK not in wb.sheetnames or SH_LOC not in wb.sheetnames:
        raise RuntimeError("Faltan catálogos CAT_STOCK_CUATI o CAT_UBICACIONES.")

    ws_cap = wb[SH_CAPTURE]
    hm_cap = header_map(ws_cap)

    stock_lookup = build_lookup_from_sheet(wb[SH_STOCK], "item_key", STOCK_HEADERS)
    loc_lookup = build_lookup_from_sheet(wb[SH_LOC], "complete_name", LOC_HEADERS)

    partner_lookup = {}

    if SH_PARTNERS in wb.sheetnames:
        partner_lookup = build_lookup_from_sheet(
            wb[SH_PARTNERS],
            "display_name",
            PARTNER_HEADERS,
        )

    required_cols = [
        "Cargar",
        "Documento origen",
        "Contacto",
        "Ubicación de origen",
        "Ubicación de destino",
        "Producto + lote disponible",
        "Cantidad a mover",
        "Comentarios",
    ]

    missing = [c for c in required_cols if c not in hm_cap]

    if missing:
        raise RuntimeError(f"Faltan columnas en CAPTURA: {missing}")

    rows_to_upload: List[List[Any]] = []
    errors: List[str] = []

    for row in range(2, ws_cap.max_row + 1):
        cargar = ws_cap.cell(row, hm_cap["Cargar"]).value

        if str(cargar or "").strip().lower() not in [
            "sí",
            "si",
            "s",
            "yes",
            "y",
            "1",
            "true",
        ]:
            continue

        doc_origin = ws_cap.cell(row, hm_cap["Documento origen"]).value
        partner_name = ws_cap.cell(row, hm_cap["Contacto"]).value
        source_name = ws_cap.cell(row, hm_cap["Ubicación de origen"]).value
        dest_name = ws_cap.cell(row, hm_cap["Ubicación de destino"]).value
        item_key = ws_cap.cell(row, hm_cap["Producto + lote disponible"]).value
        qty = ws_cap.cell(row, hm_cap["Cantidad a mover"]).value
        comments = ws_cap.cell(row, hm_cap["Comentarios"]).value

        if not any([doc_origin, partner_name, dest_name, item_key, qty]):
            continue

        row_errors = []

        if not dest_name:
            row_errors.append("falta ubicación destino")

        if not item_key:
            row_errors.append("falta producto + lote")

        if qty in (None, "") or safe_float(qty) <= 0:
            row_errors.append("cantidad inválida")

        stock = stock_lookup.get(str(item_key).strip()) if item_key else None

        if not stock:
            row_errors.append("producto+lote no encontrado en CAT_STOCK_CUATI")

        source = loc_lookup.get(str(source_name).strip()) if source_name else None

        if not source:
            row_errors.append("ubicación origen no encontrada en CAT_UBICACIONES")

        dest = loc_lookup.get(str(dest_name).strip()) if dest_name else None

        if not dest:
            row_errors.append("ubicación destino no encontrada en CAT_UBICACIONES")

        partner_id = ""

        if partner_name:
            partner = partner_lookup.get(str(partner_name).strip())
            if partner:
                partner_id = partner.get("partner_id") or ""

        if stock:
            available = safe_float(stock.get("available_quantity"))
            if safe_float(qty) > available:
                row_errors.append(
                    f"cantidad {safe_float(qty):g} mayor que disponible {available:g}"
                )

        if row_errors:
            errors.append(f"Fila {row}: " + "; ".join(row_errors))
            continue

        lot_id = stock.get("lot_id") if stock else ""

        if lot_id in (0, "0", None):
            lot_id = ""

        picking_external_id = make_odoo_external_id(
            doc_origin or "SIN_DOCUMENTO",
            partner_id or partner_name or "SIN_CONTACTO",
            source.get("location_id"),
            dest.get("location_id"),
        )

        rows_to_upload.append([
            picking_external_id,
            doc_origin or "",
            partner_id,
            source.get("location_id"),
            dest.get("location_id"),
            stock.get("product_id"),
            lot_id,
            safe_float(qty),
            source.get("location_id"),
            dest.get("location_id"),
            comments or "",
        ])

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = SH_IMPORT

    write_row(ws, 1, ODOO_IMPORT_HEADERS)

    for idx, row_values in enumerate(rows_to_upload, start=2):
        write_row(ws, idx, row_values)

    apply_basic_style(
        ws,
        max_row=max(2, len(rows_to_upload) + 1),
        max_col=len(ODOO_IMPORT_HEADERS),
    )

    set_widths(ws, {
        "A": 38,
        "B": 24,
        "C": 28,
        "D": 34,
        "E": 36,
        "F": 34,
        "G": 42,
        "H": 18,
        "I": 34,
        "J": 28,
        "K": 34,
    })

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(2, len(rows_to_upload) + 1)}"

    ws_err = out_wb.create_sheet("VALIDACION")
    ws_err["A1"] = "Resultado"
    ws_err["B1"] = "Detalle"
    ws_err["A1"].fill = HEADER_FILL
    ws_err["A1"].font = HEADER_FONT
    ws_err["B1"].fill = HEADER_FILL
    ws_err["B1"].font = HEADER_FONT

    if errors:
        ws_err["A2"] = "ERRORES"
        ws_err["B2"] = (
            "No se incluyeron las siguientes filas en IMPORT_ODOO_IDS. "
            "Corrige CAPTURA y vuelve a ejecutar preparar."
        )

        for i, e in enumerate(errors, start=3):
            ws_err.cell(i, 1, "Error")
            ws_err.cell(i, 2, e)

    else:
        unique_pickings = len({r[0] for r in rows_to_upload})
        ws_err["A2"] = "OK"
        ws_err["B2"] = (
            f"Filas listas para importar: {len(rows_to_upload)} | "
            f"Traslados agrupados: {unique_pickings}"
        )

    set_widths(ws_err, {"A": 18, "B": 130})

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = output_dir / f"SUBIR_ODOO_Traslados_Internos_IDs_v4_{timestamp}.xlsx"

    out_wb.save(output_path)

    print("\nArchivo listo para subir generado:")
    print(output_path)
    print(f"Filas listas para importar: {len(rows_to_upload)}")
    print(f"Traslados agrupados por ID externo: {len({r[0] for r in rows_to_upload})}")

    if errors:
        print("\nAdvertencia: algunas filas tuvieron errores y no fueron incluidas:")
        for e in errors[:20]:
            print("-", e)

        if len(errors) > 20:
            print(f"... y {len(errors) - 20} errores más. Revisa VALIDACION.")

    return output_path


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera plantilla de traslados internos Odoo por IDs."
    )

    sub = parser.add_subparsers(dest="command")

    p_gen = sub.add_parser(
        "generar",
        help="Conecta a Odoo y genera la plantilla de captura.",
    )

    p_gen.add_argument(
        "--origen",
        default=os.getenv("SOURCE_LOCATION_NAME", DEFAULT_SOURCE_LOCATION_NAME),
        help=f"Ubicación origen principal. Default: {DEFAULT_SOURCE_LOCATION_NAME}",
    )

    p_gen.add_argument(
        "--salida",
        default=os.getenv("OUTPUT_DIR", str(DEFAULT_OUTPUT_DIR)),
        help="Carpeta donde se guardará la plantilla.",
    )

    p_prep = sub.add_parser(
        "preparar",
        help="Lee una plantilla llena y genera archivo listo para Odoo.",
    )

    p_prep.add_argument(
        "archivo",
        help="Ruta del archivo Plantilla_Traslados_Internos_Odoo_IDs_v4_*.xlsx ya llenado.",
    )

    p_prep.add_argument(
        "--salida",
        default=None,
        help="Carpeta donde se guardará el archivo SUBIR_ODOO.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    try:
        if args.command is None:
            source_location_name = os.getenv(
                "SOURCE_LOCATION_NAME",
                DEFAULT_SOURCE_LOCATION_NAME,
            )
            output_dir = Path(
                os.getenv("OUTPUT_DIR", str(DEFAULT_OUTPUT_DIR))
            ).expanduser()
            generate_template(
                source_location_name=source_location_name,
                output_dir=output_dir,
            )

        elif args.command == "generar":
            generate_template(
                source_location_name=args.origen,
                output_dir=Path(args.salida).expanduser(),
            )

        elif args.command == "preparar":
            template_path = Path(args.archivo).expanduser()
            output_dir = Path(args.salida).expanduser() if args.salida else None
            prepare_upload_file(template_path, output_dir)

        else:
            raise RuntimeError("Comando no reconocido. Usa 'generar' o 'preparar'.")

    except Exception as exc:
        print("\nERROR:")
        print(exc)
        sys.exit(1)


if __name__ == "__main__":
    main()
